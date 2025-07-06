"""
Gerencia a lógica de execução de campanhas, incluindo o carregamento de
contatos, personalização de mensagens, envio e geração de relatórios.
"""

import re
import time
from datetime import datetime
from pathlib import Path
from typing import TYPE_CHECKING, Any

from zap_agil.utils.logging_config import logging

if TYPE_CHECKING:
    from zap_agil.core.bot_service import BotService

import openpyxl

from .constants import (
    APP_NAME,
    AUDIO_EXTENSIONS,
    IMAGE_VIDEO_EXTENSIONS,
    THEME_COLORS,
    CampaignState,
    SourceType,
)


class CampaignManager:
    """
    Gerencia a lógica de execução de campanhas no Zap Fácil.
    """

    def __init__(self, bot_service: "BotService"):
        logging.debug("Inicializando CampaignManager.")
        self.bot_service = bot_service
        self._manual_contacts: list[dict[str, Any]] = []

    def add_manual_contact(self, identifier: str) -> bool:
        """
        Adiciona um contato manualmente à lista interna.
        Retorna True se adicionado, False se duplicado.
        """
        identifier = identifier.strip()
        if not identifier:
            logging.warning("Tentativa de adicionar contato manual vazio.")
            return False
        if any(c.get("identifier", "") == identifier for c in self._manual_contacts):
            logging.info(f"Contato manual duplicado ignorado: {identifier}")
            return False
        self._manual_contacts.append({"identifier": identifier})
        logging.debug(f"Contato manual adicionado: {identifier}")
        return True

    def clear_manual_contacts(self) -> None:
        """Limpa a lista de contatos manuais."""
        logging.debug("Limpando lista de contatos manuais.")
        self._manual_contacts.clear()

    def run_campaign(self, campaign_config: dict[str, Any]) -> None:
        """
        Orquestra a execução completa de uma campanha, desde o carregamento de
        contatos até a geração do relatório final.
        """
        logging.info("Iniciando execução de campanha.")
        try:
            self.bot_service.campaign_state = CampaignState.RUNNING
            self.bot_service.notify("campaign_status", CampaignState.RUNNING)

            self.bot_service.notify(
                "log", "--- INICIANDO CAMPANHA ---", THEME_COLORS["accent_yellow"]
            )

            start_time = datetime.now()
            # Se o tipo for MANUAL_LIST e não houver contact_source, usa os contatos manuais
            if campaign_config.get(
                "source_type", "MANUAL_LIST"
            ) == "MANUAL_LIST" and not campaign_config.get("contact_source"):
                logging.debug("Usando contatos manuais para campanha.")
                campaign_config["contact_source"] = self._manual_contacts.copy()
            contacts, error = self._get_contacts_for_campaign(campaign_config)

            if error:
                logging.error(f"Erro ao obter contatos para campanha: {error}")
                self.bot_service.notify("log", error, THEME_COLORS["log_error"])
                return

            if not contacts:
                logging.warning("Nenhum contato válido encontrado para a campanha.")
                self.bot_service.notify(
                    "log",
                    "Nenhum contato válido encontrado para a campanha.",
                    THEME_COLORS["log_warning"],
                )
                return

            # Remove duplicados, mas mantém todos os campos do contato (Nome, colunas, etc)
            seen = set()
            filtered_contacts = []
            for contact in contacts:
                identifier = str(contact.get("identifier", "")).strip()
                if not identifier:
                    logging.warning("Linha vazia ignorada na lista de contatos/grupos.")
                    self.bot_service.notify(
                        "log",
                        "Linha vazia ignorada na lista de contatos/grupos.",
                        THEME_COLORS["log_warning"],
                    )
                    continue
                if identifier in seen:
                    logging.info(f"Contato/Grupo duplicado ignorado: {identifier}")
                    self.bot_service.notify(
                        "log",
                        f"Contato/Grupo duplicado '{identifier}' foi ignorado.",
                        "gray",
                    )
                    continue
                seen.add(identifier)
                # Mantém todos os campos do contato, não só identifier
                contact_copy = dict(contact)
                contact_copy["identifier"] = identifier
                filtered_contacts.append(contact_copy)
            contacts = filtered_contacts

            total = len(contacts)
            self.bot_service.notify(
                "progress_update",
                0,
                total,
                f"Iniciando campanha com {total} contatos...",
            )
            logging.info(f"Campanha iniciada com {total} contatos.")

            report_data, success_count, fail_count = [], 0, 0

            for i, contact_info in enumerate(contacts):
                # O loop principal verifica o estado da campanha a cada iteração
                if not self._handle_campaign_state_checks():
                    logging.info("Campanha interrompida pelo usuário.")
                    report_data.append("\nCampanha interrompida pelo usuário.")
                    break

                # Verifica a conexão do WhatsApp antes de cada envio
                if (
                    not self.bot_service.webdriver_manager.is_whatsapp_ready()
                    and not self.bot_service.webdriver_manager.handle_disconnection()
                ):
                    logging.error("Falha de conexão irrecuperável com o WhatsApp.")
                    report_data.append("\nCampanha abortada por falha de conexão com o WhatsApp.")
                    self.bot_service._notify(
                        "log",
                        "Falha de conexão irrecuperável.",
                        THEME_COLORS["log_error"],
                    )
                    break

                progress_msg = (
                    f"Enviando para {contact_info.get('identifier', 'N/A')} ({i + 1}/{total})"
                )
                self.bot_service.notify("progress_update", i + 1, total, progress_msg)
                self.bot_service.notify(
                    "log",
                    f"--- Processando {i + 1}/{total}: {contact_info.get('identifier', 'N/A')} ---",
                    THEME_COLORS["log_info"],
                )
                logging.debug(
                    f"Processando contato {i + 1}/{total}: {contact_info.get('identifier', 'N/A')}"
                )

                success, message = self._process_single_contact(contact_info, campaign_config)

                if success:
                    success_count += 1
                else:
                    fail_count += 1
                report_data.append(message)

                if i < total - 1 and self.bot_service.campaign_state == CampaignState.RUNNING:
                    delay = campaign_config.get("delay", 2.0)
                    logging.debug(f"Aguardando {delay}s para o próximo envio...")
                    self.bot_service.notify(
                        "log", f"Aguardando {delay}s para o próximo envio...", "gray"
                    )
                    time.sleep(delay)

            self._generate_report(
                report_data,
                start_time,
                datetime.now(),
                total,
                success_count,
                fail_count,
            )

        except Exception as e:
            logging.exception(f"ERRO CRÍTICO NA CAMPANHA: {e}")
            self.bot_service.notify(
                "log", f"ERRO CRÍTICO NA CAMPANHA: {e}", THEME_COLORS["log_error"]
            )
        finally:
            # Garante que, mesmo com erro, o estado final seja notificado
            if self.bot_service.campaign_state != CampaignState.STOPPED:
                self.bot_service.campaign_state = CampaignState.FINISHED
                self.bot_service.notify("campaign_status", CampaignState.FINISHED)

            self.bot_service.notify(
                "log", "--- CAMPANHA FINALIZADA ---", THEME_COLORS["accent_yellow"]
            )

    def _handle_campaign_state_checks(self) -> bool:
        """
        Verifica os estados de pausa e parada da campanha de forma eficiente.
        Retorna False se a campanha deve ser interrompida.
        """
        logging.debug("Verificando estado de pausa/parada da campanha.")
        # A thread ficará "dormindo" aqui sem consumir CPU se a campanha for pausada.
        # Ela será "acordada" quando o evento for setado (retomada ou parada).
        self.bot_service.pause_event.wait()

        # Após ser acordada, verifica se o motivo foi uma ordem de parada.
        if self.bot_service.campaign_state == CampaignState.STOPPED:
            logging.info("Campanha foi parada pelo usuário.")
            return False
        return True

    def _get_contacts_for_campaign(
        self, campaign_config: dict
    ) -> tuple[list[dict[str, Any]], str | None]:
        """Carrega contatos da fonte especificada na configuração da campanha."""
        source_type_key = campaign_config.get("source_type", "MANUAL_LIST")
        source_type = SourceType[source_type_key]
        contact_source = campaign_config.get("contact_source")
        contacts: list[dict[str, Any]] = []

        try:
            if source_type == SourceType.MANUAL_LIST and isinstance(contact_source, list):
                # A lista manual já vem pronta da UI
                contacts = contact_source
            elif source_type in [SourceType.LIST, SourceType.GROUP_LIST] and isinstance(
                contact_source, str
            ):
                file_path = Path(contact_source)
                if not file_path.exists():
                    logging.error(f"Arquivo de contatos não encontrado: {file_path.name}")
                    return [], f"Arquivo de contatos não encontrado: {file_path.name}"

                if file_path.suffix.lower() == ".txt":
                    logging.debug(f"Carregando contatos de arquivo TXT: {file_path}")
                    contacts = self._load_from_txt(file_path, source_type)
                elif file_path.suffix.lower() == ".xlsx":
                    logging.debug(f"Carregando contatos de arquivo XLSX: {file_path}")
                    contacts = self._load_from_xlsx(file_path, source_type)
                else:
                    logging.error(f"Formato de arquivo não suportado: {file_path.suffix}")
                    return [], f"Formato de arquivo não suportado: {file_path.suffix}"

            self.bot_service._notify(
                "log", f"{len(contacts)} contatos carregados para a campanha.", "gray"
            )
            logging.info(f"{len(contacts)} contatos carregados para a campanha.")
            return contacts, None

        except Exception as e:
            logging.exception(f"Falha ao processar a lista de contatos: {e}")
            return [], f"Falha ao processar a lista de contatos: {e}"

    def _load_from_txt(self, file_path: Path, source_type: SourceType) -> list[dict[str, Any]]:
        """Carrega contatos de um arquivo .txt."""
        contacts = []
        try:
            lines = file_path.read_text(encoding="utf-8").splitlines()
            for line in lines:
                if stripped_line := line.strip():
                    # Apenas carrega o identificador bruto.
                    # A formatação é responsabilidade do WhatsAppManager.
                    contacts.append({"identifier": stripped_line})
            logging.debug(f"{len(contacts)} contatos carregados do TXT {file_path.name}.")
            return contacts
        except Exception as e:
            logging.exception(f"Erro ao carregar contatos do TXT: {e}")
            return []

    def _load_from_xlsx(self, file_path: Path, source_type: SourceType) -> list[dict[str, Any]]:
        """Carrega contatos de uma planilha .xlsx."""
        contacts = []
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheet = workbook.active
            if not sheet or sheet.max_row < 2:
                logging.warning(f"Planilha {file_path.name} está vazia ou sem cabeçalho.")
                self.bot_service.notify(
                    "log",
                    "Planilha .xlsx está vazia ou não tem cabeçalho.",
                    THEME_COLORS["log_warning"],
                )
                return []

            headers = [str(cell.value).strip() for cell in sheet[1]]

            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                row_data = {
                    header: value
                    for header, value in zip(headers, row, strict=False)
                    if header and value is not None
                }
                if not row_data:
                    continue

                if source_type == SourceType.GROUP_LIST:
                    identifier = str(row[0]).strip() if row and row[0] else ""
                else:  # SourceType.LIST
                    phone_header = self._find_phone_header(headers)
                    if phone_header:
                        identifier = str(row_data.get(phone_header, "")).strip()
                    else:
                        # REFINAMENTO: Se não achar cabeçalho, usa a primeira coluna como fallback
                        identifier = str(row[0]).strip() if row and row[0] else ""
                        if row_num == 2 and not phone_header:  # Loga o aviso apenas uma vez
                            logging.warning(
                                f"Nenhum cabeçalho de telefone padrão encontrado em "
                                f"{file_path.name}. Usando a 1ª coluna como identificador."
                            )
                            self.bot_service.notify(
                                "log",
                                "Nenhum cabeçalho de telefone padrão encontrado. "
                                "Usando a 1ª coluna como identificador.",
                                THEME_COLORS["log_warning"],
                            )

                if identifier:
                    row_data["identifier"] = identifier
                    contacts.append(row_data)
                else:
                    logging.info(
                        f"Linha {row_num} da planilha ignorada (sem identificador válido)."
                    )
                    self.bot_service.notify(
                        "log",
                        f"Linha {row_num} da planilha ignorada (sem identificador válido).",
                        "gray",
                    )

            logging.debug(f"{len(contacts)} contatos carregados do XLSX {file_path.name}.")
            return contacts
        except Exception as e:
            logging.exception(f"Erro ao carregar contatos do XLSX: {e}")
            return []

    def _process_single_contact(
        self, contact_info: dict, campaign_config: dict
    ) -> tuple[bool, str]:
        """Processa o envio para um único contato ou grupo."""
        identifier = contact_info.get("identifier", "N/A")
        status_details: list[str] = []
        success = True
        try:
            is_group = (
                SourceType[campaign_config.get("source_type", "MANUAL_LIST")]
                == SourceType.GROUP_LIST
            )
            logging.debug(f"Abrindo chat para {identifier} (grupo: {is_group})")

            # Abrir a conversa se houver algo para enviar
            if (
                campaign_config.get("message") or campaign_config.get("attachment")
            ) and not self.bot_service.whatsapp_manager.open_chat(identifier, is_group):
                logging.warning(f"Contato ou grupo '{identifier}' não encontrado ou inválido.")
                raise ValueError(f"Contato ou grupo '{identifier}' não encontrado ou inválido.")

            # Enviar mensagem de texto personalizada
            if msg_template := campaign_config.get("message", "").strip():
                personalized_msg = self._personalize_message(msg_template, contact_info)
                if not self.bot_service.whatsapp_manager.send_text_message(personalized_msg):
                    logging.warning(f"Falha ao enviar mensagem de texto para {identifier}.")
                    status_details.append("Texto: Falhou")
                    success = False
                else:
                    status_details.append("Texto: OK")

            # Enviar anexo
            if attachment_path := campaign_config.get("attachment", ""):
                ext = Path(attachment_path).suffix.lower()
                is_media = ext in IMAGE_VIDEO_EXTENSIONS or ext in AUDIO_EXTENSIONS

                if not self.bot_service.whatsapp_manager.send_attachment(attachment_path, is_media):
                    logging.warning(f"Falha ao enviar anexo para {identifier}.")
                    status_details.append("Anexo: Falhou")
                    success = False
                else:
                    status_details.append("Anexo: OK")

            if not status_details:
                logging.info(f"Nenhuma ação configurada para {identifier}.")
                return (
                    True,
                    f"Destinatário: {identifier}\tStatus: SUCESSO\t"
                    "Detalhes: Nenhuma ação configurada.",
                )

            final_status = "SUCESSO" if success else "FALHA PARCIAL"
            logging.info(f"Envio para {identifier}: {final_status} - {', '.join(status_details)}")
            return (
                success,
                f"Destinatário: {identifier}\tStatus: {final_status}\t"
                f"Detalhes: {', '.join(status_details)}",
            )

        except Exception as e:
            logging.exception(f"Falha geral ao processar {identifier}: {e}")
            self.bot_service.notify(
                "log", f"FALHA com {identifier}: {e}", THEME_COLORS["log_error"]
            )
            return (
                False,
                f"Destinatário: {identifier}\tStatus: FALHA GERAL\tMotivo: {e}",
            )

    def _personalize_message(self, message: str, contact_data: dict) -> str:
        """
        Substitui tags como @Nome, @Cidade, etc. por valores dos dados do contato,
        ignorando completamente tags cujo valor esteja vazio.
        Nunca substitui por número de celular ou qualquer outro fallback.
        Para grupos, se o nome do grupo estiver vazio, a tag também é removida.
        """

        def replace_tag(match):
            tag_name = match.group(1).lower()
            # Busca direta e variantes
            for key, value in contact_data.items():
                if str(key).strip().lower() == tag_name:
                    if str(value).strip():
                        return str(value)
                    else:
                        return ""  # Remove a tag se o valor estiver vazio
            # Para grupos, tenta usar o nome do grupo se existir e não estiver vazio
            if tag_name in ("nome", "grupo", "gruponome"):
                if "Nome" in contact_data and str(contact_data["Nome"]).strip():
                    return str(contact_data["Nome"])
                if "Grupo" in contact_data and str(contact_data["Grupo"]).strip():
                    return str(contact_data["Grupo"])
                return ""  # Remove a tag se não houver nome/grupo
            # Nunca retorna número do contato como fallback!
            return ""  # Remove a tag se não encontrar correspondência

        return re.sub(r"@(\w+)", replace_tag, message, flags=re.IGNORECASE)

    def _generate_report(
        self,
        report_data: list,
        start_time: datetime,
        end_time: datetime,
        total: int,
        success: int,
        fail: int,
    ) -> None:
        """Cria e salva o arquivo de relatório da campanha."""
        report_filename = (
            self.bot_service.reports_dir
            / f"Relatorio_{start_time.strftime('%Y-%m-%d_%H-%M-%S')}.txt"
        )
        duration = end_time - start_time
        header = (
            f"{'=' * 60}\n"
            f" Relatório de Campanha - {APP_NAME}\n"
            f"{'=' * 60}\n\n"
            f" Início: {start_time.strftime('%d/%m/%Y %H:%M:%S')}\n"
            f" Fim:    {end_time.strftime('%d/%m/%Y %H:%M:%S')}\n"
            f" Duração: {str(duration).split('.')[0]}\n\n"
            f" RESUMO:\n"
            f"   - Total de contatos processados: {total}\n"
            f"   - Envios com sucesso: {success}\n"
            f"   - Envios com falha: {fail}\n\n"
            f"{'=' * 60}\n"
            f" DETALHES DO ENVIO\n"
            f"{'=' * 60}\n\n"
        )
        try:
            with open(report_filename, "w", encoding="utf-8") as f:
                f.write(header)
                f.write("\n".join(report_data))
            logging.info(f"Relatório salvo: {report_filename}")
            self.bot_service.notify(
                "log",
                f"Relatório salvo: {report_filename.name}",
                THEME_COLORS["accent_purple"],
            )
        except OSError as e:
            logging.exception(f"Falha ao salvar relatório: {e}")
            self.bot_service.notify(
                "log",
                f"ERRO: Falha ao salvar relatório: {e}",
                THEME_COLORS["log_error"],
            )

    def _find_phone_header(self, headers: list[str]) -> str | None:
        """Encontra o nome da coluna de telefone em uma lista de cabeçalhos."""
        phone_col_variants = [
            "numero",
            "número",
            "telefone",
            "celular",
            "phone",
            "whatsapp",
            "contato",
        ]
        for header in headers:
            if str(header).strip().lower() in phone_col_variants:
                return header
        return None
